import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    #Loading the workbook
    wb = xl.load_workbook(filename)
    #Taking reference to the 1st sheet of the workbook
    sheet =  wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1,1)
    # print(cell.value)
    # print(sheet.max_row)
    
    # updating the price and saving to a new excel file
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_prices = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_prices
    
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(filename)